import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from datetime import datetime
from sqlalchemy import or_, and_, func

def obtener_codigo_equipo(tipo_equipo, equipo_id):
    """Obtiene el código de un equipo según su tipo e ID"""
    try:
        if tipo_equipo == 'Switch':
            equipo = Switch.query.get(equipo_id)
            return equipo.codigo if equipo else f'Switch-{equipo_id}'
        elif tipo_equipo == 'Gabinete':
            equipo = Gabinete.query.get(equipo_id)
            return equipo.codigo if equipo else f'Gabinete-{equipo_id}'
        elif tipo_equipo == 'Camara':
            equipo = Camara.query.get(equipo_id)
            return equipo.codigo if equipo else f'Camara-{equipo_id}'
        elif tipo_equipo == 'NVR':
            equipo = NVR_DVR.query.get(equipo_id)
            return equipo.codigo if equipo else f'NVR-{equipo_id}'
        else:
            return f'{tipo_equipo}-{equipo_id}'
    except:
        return f'{tipo_equipo}-{equipo_id}'

# Registrar función como filtro de Jinja
app.jinja_env.globals['obtener_codigo_equipo'] = obtener_codigo_equipo
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
import uuid

from models import db, Usuario, Ubicacion, Camara, Gabinete, Switch, Puerto_Switch, UPS, NVR_DVR, Fuente_Poder, Catalogo_Tipo_Falla, Falla, Mantenimiento, Equipo_Tecnico, Historial_Estado_Equipo, Enlace, VLAN, ConexionTopologia

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuración para subida de archivos
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo
app.config['UPLOAD_FOLDER'] = 'static/uploads/fallas'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}

# Crear carpeta de uploads si no existe
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Soportar SQLite (desarrollo) y PostgreSQL (producción)
database_url = os.environ.get('DATABASE_URL', 'sqlite:///sistema_camaras.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Debug: Información de base de datos para verificar configuración Railway
print(f"🔧 DEBUG: DATABASE_URL = {os.environ.get('DATABASE_URL', 'NO DEFINIDO')}")
print(f"🔧 DEBUG: SQLALCHEMY_DATABASE_URI = {app.config.get('SQLALCHEMY_DATABASE_URI', 'NO CONFIGURADO')}")
print(f"🔧 DEBUG: Usando SQLite = {'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', '').lower()}")

# Logging del usuario Charles si existe
try:
    from models import Usuario
    charles_check = Usuario.query.filter_by(username='charles.jelvez').first()
    if charles_check:
        print(f"✅ Charles encontrado en BD: {charles_check.username} ({charles_check.rol})")
    else:
        print("❌ Charles NO encontrado en BD")
        total_users = Usuario.query.count()
        print(f"📊 Total usuarios en BD: {total_users}")
except Exception as e:
    print(f"❌ Error verificando usuarios: {e}")
login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# Decorador para verificar roles
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.rol not in roles:
                flash('No tienes permisos para acceder a esta página', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Función para obtener modo del sistema
def obtener_modo_sistema():
    """Obtiene el modo actual del sistema (demo o real)"""
    return session.get('modo_sistema', 'real')

# Función de validación anti-duplicados
def validar_falla_duplicada(equipo_tipo, equipo_id):
    """Valida si se puede insertar una nueva falla para un equipo"""
    falla_activa = Falla.query.filter_by(
        equipo_tipo=equipo_tipo,
        equipo_id=equipo_id
    ).filter(
        Falla.estado.in_(['Pendiente', 'Asignada', 'En Proceso'])
    ).order_by(Falla.fecha_reporte.desc()).first()
    
    if falla_activa:
        return {
            'permitir': False,
            'mensaje': f'Ya existe una falla {falla_activa.estado} para este equipo (ID: {falla_activa.id}, reportada el {falla_activa.fecha_reporte.strftime("%d/%m/%Y")}). Debe cerrar o cancelar la falla anterior antes de reportar una nueva.',
            'falla_existente': falla_activa
        }
    
    return {
        'permitir': True,
        'mensaje': 'OK',
        'falla_existente': None
    }

# ========== AUTENTICACIÓN ==========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = Usuario.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.activo:
            login_user(user)
            flash(f'Bienvenido {user.nombre_completo}', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada correctamente', 'info')
    return redirect(url_for('login'))

# ========== DASHBOARD ==========
@app.route('/')
@login_required
def dashboard():
    # Estadísticas
    total_camaras = Camara.query.count()
    camaras_activas = Camara.query.filter_by(estado='Activo').count()
    fallas_pendientes = Falla.query.filter_by(estado='Pendiente').count()
    fallas_asignadas = Falla.query.filter_by(estado='Asignada').count()
    fallas_en_proceso = Falla.query.filter_by(estado='En Proceso').count()
    mantenimientos_mes = Mantenimiento.query.filter(
        func.extract('month', Mantenimiento.fecha) == datetime.now().month,
        func.extract('year', Mantenimiento.fecha) == datetime.now().year
    ).count()
    
    # Últimas fallas
    ultimas_fallas = Falla.query.order_by(Falla.fecha_reporte.desc()).limit(10).all()
    
    return render_template('dashboard.html',
                         total_camaras=total_camaras,
                         camaras_activas=camaras_activas,
                         fallas_pendientes=fallas_pendientes,
                         fallas_asignadas=fallas_asignadas,
                         fallas_en_proceso=fallas_en_proceso,
                         mantenimientos_mes=mantenimientos_mes,
                         ultimas_fallas=ultimas_fallas)

# ========== CÁMARAS ==========
@app.route('/camaras')
@login_required
def camaras_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    busqueda = request.args.get('busqueda', '')
    
    query = Camara.query.join(Ubicacion)
    
    if campus:
        query = query.filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(Camara.estado == estado)
    if busqueda:
        query = query.filter(or_(
            Camara.codigo.like(f'%{busqueda}%'),
            Camara.nombre.like(f'%{busqueda}%'),
            Camara.ip.like(f'%{busqueda}%')
        ))
    
    camaras = query.all()
    ubicaciones = Ubicacion.query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('camaras_list.html', 
                         camaras=camaras, 
                         ubicaciones=ubicaciones,
                         campus_list=[c[0] for c in campus_list])

@app.route('/camaras/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def camaras_nuevo():
    if request.method == 'POST':
        camara = Camara(
            codigo=request.form.get('codigo'),
            nombre=request.form.get('nombre'),
            ip=request.form.get('ip'),
            modelo=request.form.get('modelo'),
            fabricante=request.form.get('fabricante'),
            tipo_camara=request.form.get('tipo_camara'),
            ubicacion_id=request.form.get('ubicacion_id'),
            gabinete_id=request.form.get('gabinete_id') or None,
            switch_id=request.form.get('switch_id') or None,
            nvr_id=request.form.get('nvr_id') or None,
            estado=request.form.get('estado', 'Activo'),
            fecha_alta=datetime.strptime(request.form.get('fecha_alta'), '%Y-%m-%d').date() if request.form.get('fecha_alta') else None,
            latitud=float(request.form.get('latitud')) if request.form.get('latitud') else None,
            longitud=float(request.form.get('longitud')) if request.form.get('longitud') else None,
            # PRIORIDAD 2: Campos de firmware
            version_firmware=request.form.get('version_firmware'),
            fecha_actualizacion_firmware=datetime.strptime(request.form.get('fecha_actualizacion_firmware'), '%Y-%m-%d').date() if request.form.get('fecha_actualizacion_firmware') else None,
            proxima_revision_firmware=datetime.strptime(request.form.get('proxima_revision_firmware'), '%Y-%m-%d').date() if request.form.get('proxima_revision_firmware') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(camara)
        db.session.commit()
        flash('Cámara creada exitosamente', 'success')
        return redirect(url_for('camaras_list'))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    switches = Switch.query.filter_by(estado='Activo').all()
    nvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    
    return render_template('camaras_form.html', 
                         ubicaciones=ubicaciones,
                         gabinetes=gabinetes,
                         switches=switches,
                         nvrs=nvrs,
                         camara=None)

@app.route('/camaras/<int:id>')
@login_required
def camaras_detalle(id):
    camara = Camara.query.get_or_404(id)
    fallas = Falla.query.filter_by(equipo_tipo='Camara', equipo_id=id).order_by(Falla.fecha_reporte.desc()).all()
    mantenimientos = Mantenimiento.query.filter_by(equipo_tipo='Camara', equipo_id=id).order_by(Mantenimiento.fecha.desc()).all()
    historial = Historial_Estado_Equipo.query.filter_by(equipo_tipo='Camara', equipo_id=id).order_by(Historial_Estado_Equipo.fecha_cambio.desc()).all()
    
    return render_template('camaras_detalle.html', 
                         camara=camara,
                         fallas=fallas,
                         mantenimientos=mantenimientos,
                         historial=historial)

@app.route('/camaras/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def camaras_editar(id):
    camara = Camara.query.get_or_404(id)
    
    if request.method == 'POST':
        estado_anterior = camara.estado
        
        camara.codigo = request.form.get('codigo')
        camara.nombre = request.form.get('nombre')
        camara.ip = request.form.get('ip')
        camara.modelo = request.form.get('modelo')
        camara.fabricante = request.form.get('fabricante')
        camara.tipo_camara = request.form.get('tipo_camara')
        camara.ubicacion_id = request.form.get('ubicacion_id')
        camara.estado = request.form.get('estado')
        camara.latitud = float(request.form.get('latitud')) if request.form.get('latitud') else None
        camara.longitud = float(request.form.get('longitud')) if request.form.get('longitud') else None
        # PRIORIDAD 2: Campos de firmware
        camara.version_firmware = request.form.get('version_firmware')
        camara.fecha_actualizacion_firmware = datetime.strptime(request.form.get('fecha_actualizacion_firmware'), '%Y-%m-%d').date() if request.form.get('fecha_actualizacion_firmware') else None
        camara.proxima_revision_firmware = datetime.strptime(request.form.get('proxima_revision_firmware'), '%Y-%m-%d').date() if request.form.get('proxima_revision_firmware') else None
        camara.observaciones = request.form.get('observaciones')
        
        # Registrar cambio de estado
        if estado_anterior != camara.estado:
            historial = Historial_Estado_Equipo(
                equipo_tipo='Camara',
                equipo_id=id,
                estado_anterior=estado_anterior,
                estado_nuevo=camara.estado,
                motivo=request.form.get('motivo_cambio', ''),
                usuario_id=current_user.id
            )
            db.session.add(historial)
        
        db.session.commit()
        flash('Cámara actualizada exitosamente', 'success')
        return redirect(url_for('camaras_detalle', id=id))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    switches = Switch.query.filter_by(estado='Activo').all()
    nvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    
    return render_template('camaras_form.html', 
                         camara=camara,
                         ubicaciones=ubicaciones,
                         gabinetes=gabinetes,
                         switches=switches,
                         nvrs=nvrs)

# ========== GABINETES ==========
@app.route('/gabinetes')
@login_required
def gabinetes_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    
    query = Gabinete.query.join(Ubicacion)
    
    if campus:
        query = query.filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(Gabinete.estado == estado)
    
    gabinetes = query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('gabinetes_list.html', 
                         gabinetes=gabinetes,
                         campus_list=[c[0] for c in campus_list])

@app.route('/gabinetes/<int:id>/mantencion')
@login_required
def gabinetes_mantencion(id):
    """Vista especial de mantención de gabinetes"""
    gabinete = Gabinete.query.get_or_404(id)
    
    # Obtener todos los equipos contenidos en este gabinete
    switches = Switch.query.filter_by(gabinete_id=id).all()
    nvrs = NVR_DVR.query.filter_by(gabinete_id=id).all()
    ups_list = UPS.query.filter_by(gabinete_id=id).all()
    fuentes = Fuente_Poder.query.filter_by(gabinete_id=id).all()
    
    # Historial de mantenimientos del gabinete
    mantenimientos = Mantenimiento.query.filter_by(
        equipo_tipo='Gabinete', 
        equipo_id=id
    ).order_by(Mantenimiento.fecha.desc()).all()
    
    return render_template('gabinetes_mantencion.html',
                         gabinete=gabinete,
                         switches=switches,
                         nvrs=nvrs,
                         ups_list=ups_list,
                         fuentes=fuentes,
                         mantenimientos=mantenimientos)

# ========== FALLAS ==========
@app.route('/fallas')
@login_required
def fallas_list():
    estado = request.args.get('estado', '')
    campus = request.args.get('campus', '')
    
    query = Falla.query
    
    if estado:
        query = query.filter(Falla.estado == estado)
    
    if current_user.rol == 'tecnico':
        query = query.filter(Falla.tecnico_asignado_id == current_user.id)
    
    fallas = query.order_by(Falla.fecha_reporte.desc()).all()
    
    return render_template('fallas_list.html', fallas=fallas)

@app.route('/fallas/nueva', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor', 'tecnico')
def fallas_nueva():
    if request.method == 'POST':
        equipo_tipo = request.form.get('equipo_tipo')
        equipo_id = int(request.form.get('equipo_id'))
        
        # Validar anti-duplicados
        validacion = validar_falla_duplicada(equipo_tipo, equipo_id)
        if not validacion['permitir']:
            flash(validacion['mensaje'], 'warning')
            return redirect(url_for('fallas_nueva'))
        
        falla = Falla(
            equipo_tipo=equipo_tipo,
            equipo_id=equipo_id,
            tipo_falla_id=request.form.get('tipo_falla_id'),
            descripcion=request.form.get('descripcion'),
            prioridad=request.form.get('prioridad'),
            reportado_por_id=current_user.id,
            estado='Pendiente'
        )
        db.session.add(falla)
        db.session.commit()
        flash('Falla reportada exitosamente', 'success')
        return redirect(url_for('fallas_list'))
    
    tipos_falla = Catalogo_Tipo_Falla.query.all()
    camaras = Camara.query.filter_by(estado='Activo').all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    switches = Switch.query.filter_by(estado='Activo').all()
    
    return render_template('fallas_form.html',
                         tipos_falla=tipos_falla,
                         camaras=camaras,
                         gabinetes=gabinetes,
                         switches=switches)

@app.route('/fallas/<int:id>')
@login_required
def fallas_detalle(id):
    falla = Falla.query.get_or_404(id)
    tecnicos = Usuario.query.filter_by(rol='tecnico', activo=True).all()
    return render_template('fallas_detalle.html', falla=falla, tecnicos=tecnicos)

@app.route('/fallas/<int:id>/asignar', methods=['POST'])
@login_required
@role_required('admin', 'supervisor')
def fallas_asignar(id):
    falla = Falla.query.get_or_404(id)
    falla.tecnico_asignado_id = request.form.get('tecnico_id')
    falla.estado = 'Asignada'
    falla.fecha_asignacion = datetime.now()
    db.session.commit()
    flash('Falla asignada correctamente', 'success')
    return redirect(url_for('fallas_detalle', id=id))

@app.route('/fallas/<int:id>/iniciar', methods=['POST'])
@login_required
def fallas_iniciar(id):
    falla = Falla.query.get_or_404(id)
    if falla.tecnico_asignado_id != current_user.id and current_user.rol not in ['admin', 'supervisor']:
        flash('No tienes permisos para iniciar esta falla', 'danger')
        return redirect(url_for('fallas_detalle', id=id))
    
    falla.estado = 'En Proceso'
    falla.fecha_inicio_reparacion = datetime.now()
    db.session.commit()
    flash('Reparación iniciada', 'success')
    return redirect(url_for('fallas_detalle', id=id))

@app.route('/fallas/<int:id>/reparar', methods=['GET', 'POST'])
@login_required
def fallas_reparar(id):
    falla = Falla.query.get_or_404(id)
    
    if request.method == 'POST':
        falla.estado = 'Reparada'
        falla.fecha_fin_reparacion = datetime.now()
        falla.solucion_aplicada = request.form.get('solucion_aplicada')
        falla.materiales_utilizados = request.form.get('materiales_utilizados')
        falla.costo_reparacion = float(request.form.get('costo_reparacion', 0))
        
        # Manejar subida de fotos
        fotos_uploaded = []
        if 'fotos' in request.files:
            for file in request.files.getlist('fotos'):
                if file and file.filename and allowed_file(file.filename):
                    # Generar nombre único para el archivo
                    original_filename = secure_filename(file.filename)
                    extension = original_filename.rsplit('.', 1)[1].lower()
                    unique_filename = f"falla_{id}_{uuid.uuid4().hex}.{extension}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    fotos_uploaded.append(f"/static/uploads/fallas/{unique_filename}")
        
        # Actualizar fotos existentes o agregar las nuevas
        if fotos_uploaded:
            if falla.fotos_reparacion:
                # Agregar a las fotos existentes
                fotos_existentes = falla.fotos_reparacion.split(',')
                falla.fotos_reparacion = ','.join(fotos_existentes + fotos_uploaded)
            else:
                # Primera foto
                falla.fotos_reparacion = ','.join(fotos_uploaded)
        
        if falla.fecha_inicio_reparacion:
            delta = falla.fecha_fin_reparacion - falla.fecha_inicio_reparacion
            falla.tiempo_resolucion_horas = delta.total_seconds() / 3600
        
        db.session.commit()
        flash('Falla marcada como reparada' + (' con fotos subidas' if fotos_uploaded else ''), 'success')
        return redirect(url_for('fallas_detalle', id=id))
    
    return render_template('fallas_reparar.html', falla=falla)

@app.route('/fallas/<int:id>/eliminar-foto', methods=['POST'])
@login_required
def eliminar_foto_reparacion(id):
    falla = Falla.query.get_or_404(id)
    foto_url = request.form.get('foto_url')
    
    if falla.fotos_reparacion and foto_url:
        fotos_list = falla.fotos_reparacion.split(',')
        if foto_url in fotos_list:
            fotos_list.remove(foto_url)
            falla.fotos_reparacion = ','.join(fotos_list) if fotos_list else None
            
            # Eliminar archivo físico
            photo_path = os.path.join(os.getcwd(), foto_url.lstrip('/'))
            if os.path.exists(photo_path):
                os.remove(photo_path)
            
            db.session.commit()
            flash('Foto eliminada exitosamente', 'success')
    
    return redirect(url_for('fallas_reparar', id=id))

@app.route('/fallas/<int:id>/cerrar', methods=['POST'])
@login_required
@role_required('admin', 'supervisor')
def fallas_cerrar(id):
    falla = Falla.query.get_or_404(id)
    falla.estado = 'Cerrada'
    falla.fecha_cierre = datetime.now()
    db.session.commit()
    flash('Falla cerrada correctamente', 'success')
    return redirect(url_for('fallas_detalle', id=id))

# ========== MANTENIMIENTOS ==========
@app.route('/mantenimientos')
@login_required
def mantenimientos_list():
    mantenimientos = Mantenimiento.query.order_by(Mantenimiento.fecha.desc()).all()
    return render_template('mantenimientos_list.html', mantenimientos=mantenimientos)

@app.route('/mantenimientos/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor', 'tecnico')
def mantenimientos_nuevo():
    if request.method == 'POST':
        mantenimiento = Mantenimiento(
            equipo_tipo=request.form.get('equipo_tipo'),
            equipo_id=request.form.get('equipo_id'),
            tipo=request.form.get('tipo'),
            fecha=datetime.now(),
            tecnico_id=current_user.id,
            descripcion=request.form.get('descripcion'),
            materiales_utilizados=request.form.get('materiales_utilizados'),
            tiempo_ejecucion_horas=float(request.form.get('tiempo_ejecucion_horas', 0)),
            costo=float(request.form.get('costo', 0)),
            observaciones=request.form.get('observaciones')
        )
        db.session.add(mantenimiento)
        db.session.commit()
        flash('Mantenimiento registrado exitosamente', 'success')
        return redirect(url_for('mantenimientos_list'))
    
    return render_template('mantenimientos_form.html')

# ========== MAPAS ==========
@app.route('/mapa-red')
@login_required
def mapa_red():
    campus = request.args.get('campus', '')
    
    # Obtener datos para generar diagrama Mermaid
    if campus:
        ubicaciones = Ubicacion.query.filter_by(campus=campus).all()
    else:
        ubicaciones = Ubicacion.query.all()
    
    gabinetes = Gabinete.query.join(Ubicacion).filter(Ubicacion.id.in_([u.id for u in ubicaciones])).all()
    switches = Switch.query.filter(Switch.gabinete_id.in_([g.id for g in gabinetes])).all()
    
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('mapa_red.html',
                         campus_list=[c[0] for c in campus_list],
                         gabinetes=gabinetes,
                         switches=switches)

@app.route('/mapa-geolocalizacion')
@login_required
def mapa_geolocalizacion():
    camaras = Camara.query.filter(Camara.latitud.isnot(None), Camara.longitud.isnot(None)).all()
    gabinetes = Gabinete.query.filter(Gabinete.latitud.isnot(None), Gabinete.longitud.isnot(None)).all()
    
    return render_template('mapa_geolocalizacion.html',
                         camaras=camaras,
                         gabinetes=gabinetes)

@app.route('/mapa-gabinetes-gps')
@login_required
def mapa_gabinetes_gps():
    """Mapa específico para localizar gabinetes por coordenadas GPS"""
    gabinetes = Gabinete.query.filter(
        Gabinete.latitud.isnot(None),
        Gabinete.longitud.isnot(None),
        Gabinete.estado == 'Activo'
    ).all()
    
    # Preparar datos para el mapa
    gabinetes_data = []
    for gabinete in gabinetes:
        gabinete_data = {
            'id': gabinete.id,
            'codigo': gabinete.codigo,
            'nombre': gabinete.nombre or '',
            'lat': gabinete.latitud,
            'lng': gabinete.longitud,
            'ubicacion': '',
            'estado': gabinete.estado,
            'switches': len(gabinete.switches),
            'camaras': len(gabinete.camaras),
            'ups': len(gabinete.ups_list)
        }
        
        if gabinete.ubicacion:
            gabinete_data['ubicacion'] = f"{gabinete.ubicacion.campus} - {gabinete.ubicacion.edificio}"
            if gabinete.ubicacion.piso:
                gabinete_data['ubicacion'] += f" ({gabinete.ubicacion.piso})"
        
        gabinetes_data.append(gabinete_data)
    
    return render_template('mapa_gabinetes_gps.html', gabinetes=gabinetes_data)

@app.route('/informes-avanzados')
@login_required
def informes_avanzados():
    # Estadísticas generales
    stats_por_campus = db.session.query(
        Ubicacion.campus,
        func.count(Camara.id)
    ).join(Camara).group_by(Ubicacion.campus).all()
    
    fallas_por_tipo = db.session.query(
        Catalogo_Tipo_Falla.nombre,
        func.count(Falla.id)
    ).join(Falla).group_by(Catalogo_Tipo_Falla.nombre).all()
    
    return render_template('informes_avanzados.html',
                         stats_por_campus=stats_por_campus,
                         fallas_por_tipo=fallas_por_tipo)

# ========== API REST ==========
@app.route('/api/estadisticas')
@login_required
def api_estadisticas():
    stats = {
        'total_camaras': Camara.query.count(),
        'camaras_activas': Camara.query.filter_by(estado='Activo').count(),
        'fallas_pendientes': Falla.query.filter_by(estado='Pendiente').count(),
        'fallas_en_proceso': Falla.query.filter_by(estado='En Proceso').count()
    }
    return jsonify(stats)

@app.route('/api/fallas/validar')
@login_required
def api_validar_falla():
    equipo_tipo = request.args.get('equipo_tipo')
    equipo_id = request.args.get('equipo_id')
    
    if not equipo_tipo or not equipo_id:
        return jsonify({'error': 'Parámetros faltantes'}), 400
    
    resultado = validar_falla_duplicada(equipo_tipo, int(equipo_id))
    return jsonify(resultado)

@app.route('/api/gabinetes/<int:id>/equipos')
@login_required
def api_gabinete_equipos(id):
    """API para obtener todos los equipos de un gabinete"""
    gabinete = Gabinete.query.get_or_404(id)
    
    switches = Switch.query.filter_by(gabinete_id=id).all()
    nvrs = NVR_DVR.query.filter_by(gabinete_id=id).all()
    ups_list = UPS.query.filter_by(gabinete_id=id).all()
    fuentes = Fuente_Poder.query.filter_by(gabinete_id=id).all()
    
    return jsonify({
        'gabinete': {
            'id': gabinete.id,
            'codigo': gabinete.codigo,
            'nombre': gabinete.nombre
        },
        'switches': [{
            'id': s.id,
            'codigo': s.codigo,
            'modelo': s.modelo,
            'ip': s.ip,
            'estado': s.estado
        } for s in switches],
        'nvrs': [{
            'id': n.id,
            'codigo': n.codigo,
            'tipo': n.tipo,
            'modelo': n.modelo,
            'estado': n.estado
        } for n in nvrs],
        'ups': [{
            'id': u.id,
            'codigo': u.codigo,
            'modelo': u.modelo,
            'capacidad_va': u.capacidad_va,
            'estado': u.estado
        } for u in ups_list],
        'fuentes': [{
            'id': f.id,
            'codigo': f.codigo,
            'modelo': f.modelo,
            'voltaje': f.voltaje,
            'estado': f.estado
        } for f in fuentes]
    })

# ========== SWITCHES ==========
@app.route('/switches')
@login_required
def switches_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    busqueda = request.args.get('busqueda', '')
    
    query = Switch.query.join(Gabinete).join(Ubicacion)
    
    if campus:
        query = query.filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(Switch.estado == estado)
    if busqueda:
        query = query.filter(or_(
            Switch.codigo.like(f'%{busqueda}%'),
            Switch.nombre.like(f'%{busqueda}%'),
            Switch.ip.like(f'%{busqueda}%')
        ))
    
    switches = query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('switches_list.html', 
                         switches=switches,
                         campus_list=[c[0] for c in campus_list])

@app.route('/switches/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def switches_nuevo():
    if request.method == 'POST':
        switch = Switch(
            codigo=request.form.get('codigo'),
            nombre=request.form.get('nombre'),
            ip=request.form.get('ip'),
            modelo=request.form.get('modelo'),
            marca=request.form.get('marca'),
            numero_serie=request.form.get('numero_serie'),
            gabinete_id=request.form.get('gabinete_id') or None,
            puertos_totales=int(request.form.get('puertos_totales', 0)),
            puertos_usados=int(request.form.get('puertos_usados', 0)),
            puertos_disponibles=int(request.form.get('puertos_disponibles', 0)),
            capacidad_poe=request.form.get('capacidad_poe') == 'on',
            estado=request.form.get('estado', 'Activo'),
            fecha_alta=datetime.strptime(request.form.get('fecha_alta'), '%Y-%m-%d').date() if request.form.get('fecha_alta') else None,
            latitud=float(request.form.get('latitud')) if request.form.get('latitud') else None,
            longitud=float(request.form.get('longitud')) if request.form.get('longitud') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(switch)
        db.session.commit()
        flash('Switch creado exitosamente', 'success')
        return redirect(url_for('switches_list'))
    
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('switches_form.html', gabinetes=gabinetes, switch=None)

@app.route('/switches/<int:id>')
@login_required
def switches_detalle(id):
    switch = Switch.query.get_or_404(id)
    puertos = Puerto_Switch.query.filter_by(switch_id=id).all()
    camaras = Camara.query.filter_by(switch_id=id).all()
    return render_template('switches_detalle.html', 
                         switch=switch,
                         puertos=puertos,
                         camaras=camaras)

@app.route('/switches/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def switches_editar(id):
    switch = Switch.query.get_or_404(id)
    
    if request.method == 'POST':
        switch.codigo = request.form.get('codigo')
        switch.nombre = request.form.get('nombre')
        switch.ip = request.form.get('ip')
        switch.modelo = request.form.get('modelo')
        switch.marca = request.form.get('marca')
        switch.numero_serie = request.form.get('numero_serie')
        switch.gabinete_id = request.form.get('gabinete_id') or None
        switch.puertos_totales = int(request.form.get('puertos_totales', 0))
        switch.puertos_usados = int(request.form.get('puertos_usados', 0))
        switch.puertos_disponibles = int(request.form.get('puertos_disponibles', 0))
        switch.capacidad_poe = request.form.get('capacidad_poe') == 'on'
        switch.estado = request.form.get('estado')
        switch.latitud = float(request.form.get('latitud')) if request.form.get('latitud') else None
        switch.longitud = float(request.form.get('longitud')) if request.form.get('longitud') else None
        switch.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('Switch actualizado exitosamente', 'success')
        return redirect(url_for('switches_detalle', id=id))
    
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('switches_form.html', switch=switch, gabinetes=gabinetes)

@app.route('/switches/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def switches_eliminar(id):
    switch = Switch.query.get_or_404(id)
    db.session.delete(switch)
    db.session.commit()
    flash('Switch eliminado exitosamente', 'success')
    return redirect(url_for('switches_list'))

# ========== NVR/DVR ==========
@app.route('/nvr')
@login_required
def nvr_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    tipo = request.args.get('tipo', '')
    
    query = NVR_DVR.query
    
    if campus and NVR_DVR.ubicacion_id:
        query = query.join(Ubicacion).filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(NVR_DVR.estado == estado)
    if tipo:
        query = query.filter(NVR_DVR.tipo == tipo)
    
    nvrs = query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('nvr_list.html', 
                         nvrs=nvrs,
                         campus_list=[c[0] for c in campus_list])

@app.route('/nvr/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def nvr_nuevo():
    if request.method == 'POST':
        nvr = NVR_DVR(
            codigo=request.form.get('codigo'),
            tipo=request.form.get('tipo'),
            modelo=request.form.get('modelo'),
            marca=request.form.get('marca'),
            canales_totales=int(request.form.get('canales_totales', 0)),
            canales_usados=int(request.form.get('canales_usados', 0)),
            ip=request.form.get('ip'),
            ubicacion_id=request.form.get('ubicacion_id') or None,
            gabinete_id=request.form.get('gabinete_id') or None,
            estado=request.form.get('estado', 'Activo'),
            fecha_alta=datetime.strptime(request.form.get('fecha_alta'), '%Y-%m-%d').date() if request.form.get('fecha_alta') else None,
            latitud=float(request.form.get('latitud')) if request.form.get('latitud') else None,
            longitud=float(request.form.get('longitud')) if request.form.get('longitud') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(nvr)
        db.session.commit()
        flash('NVR/DVR creado exitosamente', 'success')
        return redirect(url_for('nvr_list'))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('nvr_form.html', ubicaciones=ubicaciones, gabinetes=gabinetes, nvr=None)

@app.route('/nvr/<int:id>')
@login_required
def nvr_detalle(id):
    nvr = NVR_DVR.query.get_or_404(id)
    camaras = Camara.query.filter_by(nvr_id=id).all()
    return render_template('nvr_detalle.html', nvr=nvr, camaras=camaras)

@app.route('/nvr/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def nvr_editar(id):
    nvr = NVR_DVR.query.get_or_404(id)
    
    if request.method == 'POST':
        nvr.codigo = request.form.get('codigo')
        nvr.tipo = request.form.get('tipo')
        nvr.modelo = request.form.get('modelo')
        nvr.marca = request.form.get('marca')
        nvr.canales_totales = int(request.form.get('canales_totales', 0))
        nvr.canales_usados = int(request.form.get('canales_usados', 0))
        nvr.ip = request.form.get('ip')
        nvr.ubicacion_id = request.form.get('ubicacion_id') or None
        nvr.gabinete_id = request.form.get('gabinete_id') or None
        nvr.estado = request.form.get('estado')
        nvr.latitud = float(request.form.get('latitud')) if request.form.get('latitud') else None
        nvr.longitud = float(request.form.get('longitud')) if request.form.get('longitud') else None
        nvr.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('NVR/DVR actualizado exitosamente', 'success')
        return redirect(url_for('nvr_detalle', id=id))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('nvr_form.html', nvr=nvr, ubicaciones=ubicaciones, gabinetes=gabinetes)

@app.route('/nvr/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def nvr_eliminar(id):
    nvr = NVR_DVR.query.get_or_404(id)
    db.session.delete(nvr)
    db.session.commit()
    flash('NVR/DVR eliminado exitosamente', 'success')
    return redirect(url_for('nvr_list'))

# ========== UPS ==========
@app.route('/ups')
@login_required
def ups_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    
    query = UPS.query
    
    if campus and UPS.ubicacion_id:
        query = query.join(Ubicacion).filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(UPS.estado == estado)
    
    ups_list = query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('ups_list.html', 
                         ups_list=ups_list,
                         campus_list=[c[0] for c in campus_list])

@app.route('/ups/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def ups_nuevo():
    if request.method == 'POST':
        ups = UPS(
            codigo=request.form.get('codigo'),
            modelo=request.form.get('modelo'),
            marca=request.form.get('marca'),
            capacidad_va=int(request.form.get('capacidad_va', 0)),
            numero_baterias=int(request.form.get('numero_baterias', 0)),
            ubicacion_id=request.form.get('ubicacion_id') or None,
            gabinete_id=request.form.get('gabinete_id') or None,
            equipos_que_alimenta=request.form.get('equipos_que_alimenta'),
            estado=request.form.get('estado', 'Activo'),
            fecha_alta=datetime.strptime(request.form.get('fecha_alta'), '%Y-%m-%d').date() if request.form.get('fecha_alta') else None,
            fecha_instalacion=datetime.strptime(request.form.get('fecha_instalacion'), '%Y-%m-%d').date() if request.form.get('fecha_instalacion') else None,
            latitud=float(request.form.get('latitud')) if request.form.get('latitud') else None,
            longitud=float(request.form.get('longitud')) if request.form.get('longitud') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(ups)
        db.session.commit()
        flash('UPS creado exitosamente', 'success')
        return redirect(url_for('ups_list'))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('ups_form.html', ubicaciones=ubicaciones, gabinetes=gabinetes, ups=None)

@app.route('/ups/<int:id>')
@login_required
def ups_detalle(id):
    ups = UPS.query.get_or_404(id)
    mantenimientos = Mantenimiento.query.filter_by(equipo_tipo='UPS', equipo_id=id).order_by(Mantenimiento.fecha.desc()).all()
    return render_template('ups_detalle.html', ups=ups, mantenimientos=mantenimientos)

@app.route('/ups/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def ups_editar(id):
    ups = UPS.query.get_or_404(id)
    
    if request.method == 'POST':
        ups.codigo = request.form.get('codigo')
        ups.modelo = request.form.get('modelo')
        ups.marca = request.form.get('marca')
        ups.capacidad_va = int(request.form.get('capacidad_va', 0))
        ups.numero_baterias = int(request.form.get('numero_baterias', 0))
        ups.ubicacion_id = request.form.get('ubicacion_id') or None
        ups.gabinete_id = request.form.get('gabinete_id') or None
        ups.equipos_que_alimenta = request.form.get('equipos_que_alimenta')
        ups.estado = request.form.get('estado')
        ups.latitud = float(request.form.get('latitud')) if request.form.get('latitud') else None
        ups.longitud = float(request.form.get('longitud')) if request.form.get('longitud') else None
        ups.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('UPS actualizado exitosamente', 'success')
        return redirect(url_for('ups_detalle', id=id))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('ups_form.html', ups=ups, ubicaciones=ubicaciones, gabinetes=gabinetes)

@app.route('/ups/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def ups_eliminar(id):
    ups = UPS.query.get_or_404(id)
    db.session.delete(ups)
    db.session.commit()
    flash('UPS eliminado exitosamente', 'success')
    return redirect(url_for('ups_list'))

# ========== FUENTES DE PODER ==========
@app.route('/fuentes')
@login_required
def fuentes_list():
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    
    query = Fuente_Poder.query
    
    if campus and Fuente_Poder.ubicacion_id:
        query = query.join(Ubicacion).filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(Fuente_Poder.estado == estado)
    
    fuentes = query.all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('fuentes_list.html', 
                         fuentes=fuentes,
                         campus_list=[c[0] for c in campus_list])

@app.route('/fuentes/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def fuentes_nuevo():
    if request.method == 'POST':
        fuente = Fuente_Poder(
            codigo=request.form.get('codigo'),
            modelo=request.form.get('modelo'),
            voltaje=request.form.get('voltaje'),
            amperaje=request.form.get('amperaje'),
            equipos_que_alimenta=request.form.get('equipos_que_alimenta'),
            ubicacion_id=request.form.get('ubicacion_id') or None,
            gabinete_id=request.form.get('gabinete_id') or None,
            estado=request.form.get('estado', 'Activo'),
            fecha_alta=datetime.strptime(request.form.get('fecha_alta'), '%Y-%m-%d').date() if request.form.get('fecha_alta') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(fuente)
        db.session.commit()
        flash('Fuente de poder creada exitosamente', 'success')
        return redirect(url_for('fuentes_list'))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('fuentes_form.html', ubicaciones=ubicaciones, gabinetes=gabinetes, fuente=None)

@app.route('/fuentes/<int:id>')
@login_required
def fuentes_detalle(id):
    fuente = Fuente_Poder.query.get_or_404(id)
    return render_template('fuentes_detalle.html', fuente=fuente)

@app.route('/fuentes/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def fuentes_editar(id):
    fuente = Fuente_Poder.query.get_or_404(id)
    
    if request.method == 'POST':
        fuente.codigo = request.form.get('codigo')
        fuente.modelo = request.form.get('modelo')
        fuente.voltaje = request.form.get('voltaje')
        fuente.amperaje = request.form.get('amperaje')
        fuente.equipos_que_alimenta = request.form.get('equipos_que_alimenta')
        fuente.ubicacion_id = request.form.get('ubicacion_id') or None
        fuente.gabinete_id = request.form.get('gabinete_id') or None
        fuente.estado = request.form.get('estado')
        fuente.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('Fuente de poder actualizada exitosamente', 'success')
        return redirect(url_for('fuentes_detalle', id=id))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    return render_template('fuentes_form.html', fuente=fuente, ubicaciones=ubicaciones, gabinetes=gabinetes)

@app.route('/fuentes/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def fuentes_eliminar(id):
    fuente = Fuente_Poder.query.get_or_404(id)
    db.session.delete(fuente)
    db.session.commit()
    flash('Fuente de poder eliminada exitosamente', 'success')
    return redirect(url_for('fuentes_list'))

# ========== PUERTOS SWITCH ==========
@app.route('/puertos')
@login_required
def puertos_list():
    switch_id = request.args.get('switch_id', '')
    estado = request.args.get('estado', '')
    
    query = Puerto_Switch.query
    
    if switch_id:
        query = query.filter(Puerto_Switch.switch_id == switch_id)
    if estado:
        query = query.filter(Puerto_Switch.estado == estado)
    
    puertos = query.all()
    switches = Switch.query.filter_by(estado='Activo').all()
    
    return render_template('puertos_list.html', 
                         puertos=puertos,
                         switches=switches)

@app.route('/puertos/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor', 'tecnico')
def puertos_nuevo():
    if request.method == 'POST':
        puerto = Puerto_Switch(
            switch_id=request.form.get('switch_id'),
            numero_puerto=int(request.form.get('numero_puerto')),
            camara_id=request.form.get('camara_id') or None,
            ip_dispositivo=request.form.get('ip_dispositivo'),
            estado=request.form.get('estado', 'Disponible'),
            tipo_conexion=request.form.get('tipo_conexion'),
            nvr_id=request.form.get('nvr_id') or None,
            puerto_nvr=request.form.get('puerto_nvr')
        )
        db.session.add(puerto)
        db.session.commit()
        flash('Puerto creado exitosamente', 'success')
        return redirect(url_for('puertos_list'))
    
    switches = Switch.query.filter_by(estado='Activo').all()
    camaras = Camara.query.filter_by(estado='Activo').all()
    nvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    return render_template('puertos_form.html', switches=switches, camaras=camaras, nvrs=nvrs, puerto=None)

@app.route('/puertos/<int:id>')
@login_required
def puertos_detalle(id):
    puerto = Puerto_Switch.query.get_or_404(id)
    return render_template('puertos_detalle.html', puerto=puerto)

@app.route('/puertos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor', 'tecnico')
def puertos_editar(id):
    puerto = Puerto_Switch.query.get_or_404(id)
    
    if request.method == 'POST':
        puerto.switch_id = request.form.get('switch_id')
        puerto.numero_puerto = int(request.form.get('numero_puerto'))
        puerto.camara_id = request.form.get('camara_id') or None
        puerto.ip_dispositivo = request.form.get('ip_dispositivo')
        puerto.estado = request.form.get('estado')
        puerto.tipo_conexion = request.form.get('tipo_conexion')
        puerto.nvr_id = request.form.get('nvr_id') or None
        puerto.puerto_nvr = request.form.get('puerto_nvr')
        
        db.session.commit()
        flash('Puerto actualizado exitosamente', 'success')
        return redirect(url_for('puertos_detalle', id=id))
    
    switches = Switch.query.filter_by(estado='Activo').all()
    camaras = Camara.query.filter_by(estado='Activo').all()
    nvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    return render_template('puertos_form.html', puerto=puerto, switches=switches, camaras=camaras, nvrs=nvrs)

@app.route('/puertos/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def puertos_eliminar(id):
    puerto = Puerto_Switch.query.get_or_404(id)
    db.session.delete(puerto)
    db.session.commit()
    flash('Puerto eliminado exitosamente', 'success')
    return redirect(url_for('puertos_list'))

# ========== EQUIPOS TECNICOS ==========
@app.route('/tecnicos')
@login_required
def tecnicos_list():
    especialidad = request.args.get('especialidad', '')
    estado = request.args.get('estado', '')
    
    query = Equipo_Tecnico.query
    
    if especialidad:
        query = query.filter(Equipo_Tecnico.especialidad == especialidad)
    if estado:
        query = query.filter(Equipo_Tecnico.estado == estado)
    
    tecnicos = query.all()
    especialidades = db.session.query(Equipo_Tecnico.especialidad).distinct().all()
    
    return render_template('tecnicos_list.html', 
                         tecnicos=tecnicos,
                         especialidades=[e[0] for e in especialidades if e[0]])

@app.route('/tecnicos/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def tecnicos_nuevo():
    if request.method == 'POST':
        tecnico = Equipo_Tecnico(
            nombre=request.form.get('nombre'),
            apellido=request.form.get('apellido'),
            especialidad=request.form.get('especialidad'),
            telefono=request.form.get('telefono'),
            email=request.form.get('email'),
            estado=request.form.get('estado', 'Activo'),
            fecha_ingreso=datetime.strptime(request.form.get('fecha_ingreso'), '%Y-%m-%d').date() if request.form.get('fecha_ingreso') else None
        )
        db.session.add(tecnico)
        db.session.commit()
        flash('Técnico creado exitosamente', 'success')
        return redirect(url_for('tecnicos_list'))
    
    return render_template('tecnicos_form.html', tecnico=None)

@app.route('/tecnicos/<int:id>')
@login_required
def tecnicos_detalle(id):
    tecnico = Equipo_Tecnico.query.get_or_404(id)
    fallas_asignadas = Falla.query.filter_by(tecnico_asignado_id=id).order_by(Falla.fecha_reporte.desc()).all()
    mantenimientos = Mantenimiento.query.filter_by(tecnico_id=id).order_by(Mantenimiento.fecha.desc()).all()
    return render_template('tecnicos_detalle.html', 
                         tecnico=tecnico,
                         fallas_asignadas=fallas_asignadas,
                         mantenimientos=mantenimientos)

@app.route('/tecnicos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'supervisor')
def tecnicos_editar(id):
    tecnico = Equipo_Tecnico.query.get_or_404(id)
    
    if request.method == 'POST':
        tecnico.nombre = request.form.get('nombre')
        tecnico.apellido = request.form.get('apellido')
        tecnico.especialidad = request.form.get('especialidad')
        tecnico.telefono = request.form.get('telefono')
        tecnico.email = request.form.get('email')
        tecnico.estado = request.form.get('estado')
        
        db.session.commit()
        flash('Técnico actualizado exitosamente', 'success')
        return redirect(url_for('tecnicos_detalle', id=id))
    
    return render_template('tecnicos_form.html', tecnico=tecnico)

@app.route('/tecnicos/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def tecnicos_eliminar(id):
    tecnico = Equipo_Tecnico.query.get_or_404(id)
    db.session.delete(tecnico)
    db.session.commit()
    flash('Técnico eliminado exitosamente', 'success')
    return redirect(url_for('tecnicos_list'))
# ========== ADMINISTRACIÓN DE USUARIOS ==========
@app.route('/admin/usuarios')
@login_required
@role_required('superadmin', 'admin')
def admin_usuarios():
    """Lista todos los usuarios del sistema"""
    usuarios = Usuario.query.order_by(Usuario.fecha_creacion.desc()).all()
    return render_template('admin_usuarios_list.html', usuarios=usuarios)

@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('superadmin', 'admin')
def admin_usuarios_nuevo():
    """Crear nuevo usuario"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        rol = request.form.get('rol')
        nombre_completo = request.form.get('nombre_completo')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        
        # Validaciones
        if Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'danger')
            return render_template('admin_usuarios_form.html')
        
        # Solo SUPERADMIN puede crear otros SUPERADMIN
        if rol == 'superadmin' and current_user.rol != 'superadmin':
            flash('Solo un SUPERADMIN puede crear otro SUPERADMIN', 'danger')
            return render_template('admin_usuarios_form.html')
        
        usuario = Usuario(
            username=username,
            rol=rol,
            nombre_completo=nombre_completo,
            email=email,
            telefono=telefono,
            activo=True
        )
        usuario.set_password(password)
        
        db.session.add(usuario)
        db.session.commit()
        flash('Usuario creado exitosamente', 'success')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin_usuarios_form.html')

@app.route('/admin/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('superadmin', 'admin')
def admin_usuarios_editar(id):
    """Editar usuario existente"""
    usuario = Usuario.query.get_or_404(id)
    
    # Un usuario no puede editarse a sí mismo ni cambiar su propio rol
    if usuario.id == current_user.id:
        flash('No puedes editar tu propio perfil desde aquí', 'warning')
        return redirect(url_for('admin_usuarios'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        rol = request.form.get('rol')
        nombre_completo = request.form.get('nombre_completo')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        activo = request.form.get('activo') == 'on'
        
        # Validaciones
        if username != usuario.username and Usuario.query.filter_by(username=username).first():
            flash('El nombre de usuario ya existe', 'danger')
            return render_template('admin_usuarios_form.html', usuario=usuario)
        
        # Solo SUPERADMIN puede asignar rol SUPERADMIN
        if rol == 'superadmin' and current_user.rol != 'superadmin':
            flash('Solo un SUPERADMIN puede asignar rol SUPERADMIN', 'danger')
            return render_template('admin_usuarios_form.html', usuario=usuario)
        
        # No permitir que el último SUPERADMIN se desactive
        if usuario.rol == 'superadmin' and not activo and Usuario.query.filter_by(rol='superadmin').count() <= 1:
            flash('No se puede desactivar el último SUPERADMIN del sistema', 'danger')
            return render_template('admin_usuarios_form.html', usuario=usuario)
        
        usuario.username = username
        usuario.rol = rol
        usuario.nombre_completo = nombre_completo
        usuario.email = email
        usuario.telefono = telefono
        usuario.activo = activo
        
        db.session.commit()
        flash('Usuario actualizado exitosamente', 'success')
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin_usuarios_form.html', usuario=usuario)

@app.route('/admin/usuarios/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('superadmin', 'admin')
def admin_usuarios_eliminar(id):
    """Eliminar usuario"""
    usuario = Usuario.query.get_or_404(id)
    
    # Un usuario no puede eliminarse a sí mismo
    if usuario.id == current_user.id:
        flash('No puedes eliminar tu propio usuario', 'danger')
        return redirect(url_for('admin_usuarios'))
    
    # No permitir eliminar el último SUPERADMIN
    if usuario.rol == 'superadmin' and Usuario.query.filter_by(rol='superadmin').count() <= 1:
        flash('No se puede eliminar el último SUPERADMIN del sistema', 'danger')
        return redirect(url_for('admin_usuarios'))
    
    db.session.delete(usuario)
    db.session.commit()
    flash('Usuario eliminado exitosamente', 'success')
    return redirect(url_for('admin_usuarios'))

# ========== CONFIGURACIÓN DEL SISTEMA ==========
@app.route('/admin/configuracion')
@login_required
@role_required('superadmin')
def admin_configuracion():
    """Panel de configuración del sistema - Solo para SUPERADMIN"""
    return render_template('admin_configuracion.html')

@app.route('/admin/configuracion/modo', methods=['POST'])
@login_required
@role_required('superadmin')
def admin_configuracion_modo():
    """Cambiar entre modo Demo y Real - Solo para SUPERADMIN"""
    modo = request.form.get('modo')
    
    if modo not in ['demo', 'real']:
        flash('Modo inválido', 'danger')
        return redirect(url_for('admin_configuracion'))
    
    # Guardar modo en sesión
    session['modo_sistema'] = modo
    
    flash(f'Sistema configurado en modo {modo.upper()}', 'success')
    return redirect(url_for('admin_configuracion'))

# Inicializar base de datos y crear usuarios por defecto
@app.cli.command()
def init_db():
    """Inicializa la base de datos y crea usuarios por defecto"""
    db.create_all()
    
    # Verificar si ya existen usuarios
    if Usuario.query.count() == 0:
        usuarios = [
            Usuario(username='charles.jelvez', rol='superadmin', nombre_completo='Charles Jélvez', email='charles.jelvez@ufro.cl', activo=True),
            Usuario(username='admin', rol='admin', nombre_completo='Administrador', activo=True),
            Usuario(username='supervisor', rol='supervisor', nombre_completo='Supervisor', activo=True),
            Usuario(username='tecnico1', rol='tecnico', nombre_completo='Técnico 1', activo=True),
            Usuario(username='visualizador', rol='visualizador', nombre_completo='Visualizador', activo=True)
        ]
        
        passwords = ['charles123', 'admin123', 'super123', 'tecnico123', 'viz123']
        
        for user, password in zip(usuarios, passwords):
            user.set_password(password)
            db.session.add(user)
        
        db.session.commit()
        print('Base de datos inicializada con usuarios por defecto')
    else:
        print('Los usuarios ya existen')

# ========== RUTAS TEMPORALES PARA INICIALIZACIÓN EN RAILWAY ==========
@app.route('/init-usuarios-railway')
def init_usuarios_temporal():
    """Ruta temporal para inicializar usuarios en Railway PostgreSQL"""
    try:
        # Limpiar usuarios existentes
        Usuario.query.delete()
        db.session.commit()
        
        # Crear usuarios
        usuarios_data = [
            ('charles.jelvez', 'charles123', 'superadmin', 'Charles Jélvez', 'charles.jelvez@ufro.cl'),
            ('admin', 'admin123', 'admin', 'Administrador', 'admin@ufro.cl'),
            ('supervisor', 'super123', 'supervisor', 'Supervisor', 'supervisor@ufro.cl'),
            ('tecnico1', 'tecnico123', 'tecnico', 'Técnico 1', 'tecnico1@ufro.cl'),
            ('visualizador', 'viz123', 'visualizador', 'Visualizador', 'visualizador@ufro.cl')
        ]
        
        for username, password, rol, nombre, email in usuarios_data:
            usuario = Usuario(
                username=username,
                rol=rol,
                nombre_completo=nombre,
                email=email,
                activo=True
            )
            usuario.set_password(password)
            db.session.add(usuario)
        
        db.session.commit()
        
        return f"""
        <h1>✅ Inicialización Exitosa</h1>
        <p><strong>Usuarios creados en Railway PostgreSQL:</strong></p>
        <ul>
        <li><strong>charles.jelvez</strong> / charles123 (SUPERADMIN)</li>
        <li><strong>admin</strong> / admin123 (ADMIN)</li>
        <li><strong>supervisor</strong> / super123 (SUPERVISOR)</li>
        <li><strong>tecnico1</strong> / tecnico123 (TÉCNICO)</li>
        <li><strong>visualizador</strong> / viz123 (VISUALIZADOR)</li>
        </ul>
        <p>🔗 <a href="/login">Ir al login</a></p>
        <p><em>Esta ruta se eliminará después de la inicialización</em></p>
        """
        
    except Exception as e:
        return f"<h1>❌ Error</h1><p>{str(e)}</p>"

@app.route('/crear-charles-superadmin')
def crear_charles_superadmin():
    """Ruta para crear específicamente Charles como SUPERADMIN"""
    try:
        # Verificar si Charles ya existe
        charles = Usuario.query.filter_by(username='charles.jelvez').first()
        
        if charles:
            # Actualizar contraseña y rol
            charles.set_password('charles123')
            charles.rol = 'superadmin'
            charles.activo = True
            db.session.commit()
            mensaje = f"✅ Charles actualizado: {charles.username} ({charles.rol})"
        else:
            # Crear nuevo Charles
            charles = Usuario(
                username='charles.jelvez',
                rol='superadmin',
                nombre_completo='Charles Jélvez',
                email='charles.jelvez@ufro.cl',
                activo=True
            )
            charles.set_password('charles123')
            db.session.add(charles)
            db.session.commit()
            mensaje = f"✅ Charles creado: {charles.username} ({charles.rol})"
        
        # Mostrar todos los usuarios
        usuarios = Usuario.query.order_by(Usuario.rol.desc()).all()
        
        usuarios_html = ""
        credenciales = {
            'charles.jelvez': 'charles123',
            'admin': 'admin123',
            'supervisor': 'super123',
            'tecnico1': 'tecnico123',
            'visualizador': 'viz123'
        }
        
        for usuario in usuarios:
            pwd = credenciales.get(usuario.username, 'N/A')
            usuarios_html += f"<tr><td>{usuario.username}</td><td>{usuario.rol}</td><td>{pwd}</td><td>{'✅' if usuario.activo else '❌'}</td></tr>"
        
        return f"""
        <h1>👑 Configuración SUPERADMIN</h1>
        <p>{mensaje}</p>
        <h2>Usuarios en el Sistema:</h2>
        <table border="1" style="border-collapse: collapse; width: 100%;">
            <tr style="background: #f0f0f0;">
                <th>Usuario</th>
                <th>Rol</th>
                <th>Contraseña</th>
                <th>Estado</th>
            </tr>
            {usuarios_html}
        </table>
        <h2>🎯 Credenciales para Charles:</h2>
        <p><strong>URL:</strong> https://gestion-camaras-ufro.up.railway.app/</p>
        <p><strong>Usuario:</strong> charles.jelvez</p>
        <p><strong>Contraseña:</strong> charles123</p>
        <p><a href="/login">🔗 Ir al login</a></p>
        """
        
    except Exception as e:
        import traceback
        return f"<h1>❌ Error</h1><p>{str(e)}</p><pre>{traceback.format_exc()}</pre>"

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)

# ========== MODIFICACIÓN MASIVA DE CÁMARAS (SUPERADMIN) ==========
@app.route('/camaras/masivo', methods=['GET', 'POST'])
@login_required
@role_required('superadmin')
def camaras_masivo():
    """Modificación masiva de cámaras - Solo Superadmin"""
    if request.method == 'POST':
        # Obtener IDs de cámaras seleccionadas
        camara_ids = request.form.getlist('camara_ids[]')
        
        if not camara_ids:
            flash('Debe seleccionar al menos una cámara', 'warning')
            return redirect(url_for('camaras_masivo'))
        
        # Obtener campos a actualizar
        actualizar_estado = request.form.get('actualizar_estado')
        actualizar_ubicacion = request.form.get('actualizar_ubicacion')
        actualizar_nvr = request.form.get('actualizar_nvr')
        actualizar_gabinete = request.form.get('actualizar_gabinete')
        actualizar_switch = request.form.get('actualizar_switch')
        
        # Valores nuevos
        nuevo_estado = request.form.get('estado')
        nueva_ubicacion_id = request.form.get('ubicacion_id')
        nuevo_nvr_id = request.form.get('nvr_id')
        nuevo_gabinete_id = request.form.get('gabinete_id')
        nuevo_switch_id = request.form.get('switch_id')
        
        # Actualizar cámaras
        camaras_actualizadas = 0
        for camara_id in camara_ids:
            camara = Camara.query.get(int(camara_id))
            if camara:
                if actualizar_estado and nuevo_estado:
                    camara.estado = nuevo_estado
                if actualizar_ubicacion and nueva_ubicacion_id:
                    camara.ubicacion_id = int(nueva_ubicacion_id)
                if actualizar_nvr and nuevo_nvr_id:
                    camara.nvr_id = int(nuevo_nvr_id) if nuevo_nvr_id else None
                if actualizar_gabinete and nuevo_gabinete_id:
                    camara.gabinete_id = int(nuevo_gabinete_id) if nuevo_gabinete_id else None
                if actualizar_switch and nuevo_switch_id:
                    camara.switch_id = int(nuevo_switch_id) if nuevo_switch_id else None
                camaras_actualizadas += 1
        
        db.session.commit()
        flash(f'Se actualizaron {camaras_actualizadas} cámaras exitosamente', 'success')
        return redirect(url_for('camaras_list'))
    
    # GET - Mostrar formulario
    campus = request.args.get('campus', '')
    estado = request.args.get('estado', '')
    
    query = Camara.query.join(Ubicacion)
    if campus:
        query = query.filter(Ubicacion.campus == campus)
    if estado:
        query = query.filter(Camara.estado == estado)
    
    camaras = query.all()
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    nvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    switches = Switch.query.filter_by(estado='Activo').all()
    campus_list = db.session.query(Ubicacion.campus).distinct().all()
    
    return render_template('camaras_masivo.html',
                         camaras=camaras,
                         ubicaciones=ubicaciones,
                         nvrs=nvrs,
                         gabinetes=gabinetes,
                         switches=switches,
                         campus_list=[c[0] for c in campus_list])

# ========== SISTEMA DE INFORMES CON EXPORTACIÓN ==========
@app.route('/informes/generar', methods=['POST'])
@login_required
def informes_generar():
    """Generar informes con exportación"""
    tipo_informe = request.form.get('tipo_informe')
    formato = request.form.get('formato', 'excel')  # excel o pdf
    
    if tipo_informe == 'camaras':
        # Informe de cámaras
        campus = request.form.get('campus', '')
        estado = request.form.get('estado', '')
        
        query = Camara.query.join(Ubicacion)
        if campus:
            query = query.filter(Ubicacion.campus == campus)
        if estado:
            query = query.filter(Camara.estado == estado)
        
        camaras = query.all()
        
        if formato == 'excel':
            # Generar Excel
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from flask import send_file
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe de Cámaras"
            
            # Encabezados
            headers = ['Código', 'Nombre', 'IP', 'Modelo', 'Tipo', 'Ubicación', 'Campus', 'Estado']
            ws.append(headers)
            
            # Estilo encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Datos
            for camara in camaras:
                ws.append([
                    camara.codigo,
                    camara.nombre,
                    camara.ip or 'N/A',
                    camara.modelo or 'N/A',
                    camara.tipo_camara or 'N/A',
                    camara.ubicacion.edificio if camara.ubicacion else 'N/A',
                    camara.ubicacion.campus if camara.ubicacion else 'N/A',
                    camara.estado
                ])
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'informe_camaras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        
        elif formato == 'pdf':
            # Generar PDF
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from flask import send_file
            import io
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title = Paragraph("Informe de Cámaras - Sistema UFRO", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Tabla de datos
            data = [['Código', 'Nombre', 'Ubicación', 'Estado']]
            for camara in camaras:
                data.append([
                    camara.codigo,
                    camara.nombre or 'N/A',
                    f"{camara.ubicacion.campus if camara.ubicacion else 'N/A'} - {camara.ubicacion.edificio if camara.ubicacion else ''}",
                    camara.estado
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'informe_camaras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
    
    elif tipo_informe == 'fallas':
        # Informe de Fallas
        estado_falla = request.form.get('estado_falla', '')
        periodo = request.form.get('periodo', 'todos')
        
        query = Falla.query
        if estado_falla:
            query = query.filter(Falla.estado == estado_falla)
        
        # Filtrar por período
        if periodo == 'mes_actual':
            inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0)
            query = query.filter(Falla.fecha_reporte >= inicio_mes)
        
        fallas = query.order_by(Falla.fecha_reporte.desc()).all()
        
        if formato == 'excel':
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe de Fallas"
            
            headers = ['ID', 'Equipo Tipo', 'Equipo ID', 'Descripción', 'Prioridad', 'Estado', 'Fecha Reporte', 'Fecha Resolución']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            
            for falla in fallas:
                ws.append([
                    falla.id,
                    falla.equipo_tipo,
                    falla.equipo_id,
                    falla.descripcion[:100] if falla.descripcion else 'N/A',
                    falla.prioridad or 'N/A',
                    falla.estado,
                    falla.fecha_reporte.strftime('%d/%m/%Y %H:%M'),
                    falla.fecha_resolucion.strftime('%d/%m/%Y %H:%M') if falla.fecha_resolucion else 'Pendiente'
                ])
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'informe_fallas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
    
    elif tipo_informe == 'mantenimientos':
        # Informe de Mantenimientos
        tipo_mant = request.form.get('tipo_mantenimiento', '')
        periodo = request.form.get('periodo', 'todos')
        
        query = Mantenimiento.query
        if tipo_mant:
            query = query.filter(Mantenimiento.tipo_mantenimiento == tipo_mant)
        
        if periodo == 'mes_actual':
            inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0)
            query = query.filter(Mantenimiento.fecha >= inicio_mes)
        
        mantenimientos = query.order_by(Mantenimiento.fecha.desc()).all()
        
        if formato == 'excel':
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe de Mantenimientos"
            
            headers = ['ID', 'Equipo Tipo', 'Equipo ID', 'Tipo', 'Descripción', 'Fecha', 'Técnico', 'Costo']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
            
            for mant in mantenimientos:
                tecnico_nombre = 'N/A'
                if mant.tecnico_id:
                    tecnico = Equipo_Tecnico.query.get(mant.tecnico_id)
                    if tecnico:
                        tecnico_nombre = f"{tecnico.nombre} {tecnico.apellido}"
                
                ws.append([
                    mant.id,
                    mant.equipo_tipo,
                    mant.equipo_id,
                    mant.tipo_mantenimiento or 'N/A',
                    mant.descripcion[:100] if mant.descripcion else 'N/A',
                    mant.fecha.strftime('%d/%m/%Y %H:%M'),
                    tecnico_nombre,
                    f"${mant.costo_estimado:,.0f}" if mant.costo_estimado else 'N/A'
                ])
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'informe_mantenimientos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
    
    flash('Tipo de informe no válido', 'danger')
    return redirect(url_for('informes_avanzados'))

# ============================================================================
# PRIORIDAD 1: RUTAS ENLACES - Gestión de Conectividad
# ============================================================================

@app.route('/enlaces')
@login_required
def enlaces():
    """Listar todos los enlaces de conectividad"""
    enlaces = Enlace.query.all()
    return render_template('enlaces_list.html', enlaces=enlaces)

@app.route('/enlaces/nuevo', methods=['GET', 'POST'])
@login_required
def enlaces_nuevo():
    """Crear nuevo enlace"""
    if request.method == 'POST':
        enlace = Enlace(
            codigo=request.form['codigo'],
            nombre=request.form.get('nombre'),
            tipo_enlace=request.form['tipo_enlace'],
            origen_ubicacion_id=request.form.get('origen_ubicacion_id') or None,
            destino_ubicacion_id=request.form.get('destino_ubicacion_id') or None,
            switch_origen_id=request.form.get('switch_origen_id') or None,
            switch_destino_id=request.form.get('switch_destino_id') or None,
            latencia_ms=float(request.form['latencia_ms']) if request.form.get('latencia_ms') else None,
            porcentaje_perdida_paquetes=float(request.form['porcentaje_perdida_paquetes']) if request.form.get('porcentaje_perdida_paquetes') else None,
            estado_conexion=request.form.get('estado_conexion', 'Activo'),
            ancho_banda_mbps=int(request.form['ancho_banda_mbps']) if request.form.get('ancho_banda_mbps') else None,
            proveedor=request.form.get('proveedor'),
            fecha_instalacion=datetime.strptime(request.form['fecha_instalacion'], '%Y-%m-%d').date() if request.form.get('fecha_instalacion') else None,
            fecha_ultimo_testeo=datetime.strptime(request.form['fecha_ultimo_testeo'], '%Y-%m-%d').date() if request.form.get('fecha_ultimo_testeo') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(enlace)
        db.session.commit()
        flash('Enlace creado exitosamente', 'success')
        return redirect(url_for('enlaces'))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    switches = Switch.query.filter_by(estado='Activo').all()
    return render_template('enlaces_form.html', enlace=None, ubicaciones=ubicaciones, switches=switches)

@app.route('/enlaces/<int:id>')
@login_required
def enlaces_detalle(id):
    """Ver detalle de un enlace"""
    enlace = Enlace.query.get_or_404(id)
    return render_template('enlaces_detalle.html', enlace=enlace)

@app.route('/enlaces/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def enlaces_editar(id):
    """Editar un enlace existente"""
    enlace = Enlace.query.get_or_404(id)
    
    if request.method == 'POST':
        enlace.codigo = request.form['codigo']
        enlace.nombre = request.form.get('nombre')
        enlace.tipo_enlace = request.form['tipo_enlace']
        enlace.origen_ubicacion_id = request.form.get('origen_ubicacion_id') or None
        enlace.destino_ubicacion_id = request.form.get('destino_ubicacion_id') or None
        enlace.switch_origen_id = request.form.get('switch_origen_id') or None
        enlace.switch_destino_id = request.form.get('switch_destino_id') or None
        enlace.latencia_ms = float(request.form['latencia_ms']) if request.form.get('latencia_ms') else None
        enlace.porcentaje_perdida_paquetes = float(request.form['porcentaje_perdida_paquetes']) if request.form.get('porcentaje_perdida_paquetes') else None
        enlace.estado_conexion = request.form.get('estado_conexion', 'Activo')
        enlace.ancho_banda_mbps = int(request.form['ancho_banda_mbps']) if request.form.get('ancho_banda_mbps') else None
        enlace.proveedor = request.form.get('proveedor')
        enlace.fecha_instalacion = datetime.strptime(request.form['fecha_instalacion'], '%Y-%m-%d').date() if request.form.get('fecha_instalacion') else None
        enlace.fecha_ultimo_testeo = datetime.strptime(request.form['fecha_ultimo_testeo'], '%Y-%m-%d').date() if request.form.get('fecha_ultimo_testeo') else None
        enlace.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('Enlace actualizado exitosamente', 'success')
        return redirect(url_for('enlaces_detalle', id=id))
    
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    switches = Switch.query.filter_by(estado='Activo').all()
    return render_template('enlaces_form.html', enlace=enlace, ubicaciones=ubicaciones, switches=switches)

@app.route('/enlaces/<int:id>/eliminar', methods=['POST'])
@login_required
def enlaces_eliminar(id):
    """Eliminar un enlace"""
    enlace = Enlace.query.get_or_404(id)
    db.session.delete(enlace)
    db.session.commit()
    flash('Enlace eliminado exitosamente', 'success')
    return redirect(url_for('enlaces'))

# ============================================================================
# PRIORIDAD 3: RUTAS VLAN - Gestión de Redes Virtuales
# ============================================================================

@app.route('/vlans')
@login_required
def vlans():
    """Listar todas las VLANs"""
    vlans = VLAN.query.all()
    return render_template('vlans_list.html', vlans=vlans)

@app.route('/vlans/nuevo', methods=['GET', 'POST'])
@login_required
def vlans_nuevo():
    """Crear nueva VLAN"""
    if request.method == 'POST':
        vlan = VLAN(
            vlan_id=int(request.form['vlan_id']),
            vlan_nombre=request.form['vlan_nombre'],
            vlan_descripcion=request.form.get('vlan_descripcion'),
            red=request.form.get('red'),
            mascara=request.form.get('mascara'),
            gateway=request.form.get('gateway'),
            switch_id=request.form.get('switch_id') or None,
            estado=request.form.get('estado', 'Activo'),
            fecha_creacion=datetime.strptime(request.form['fecha_creacion'], '%Y-%m-%d').date() if request.form.get('fecha_creacion') else None,
            observaciones=request.form.get('observaciones')
        )
        db.session.add(vlan)
        db.session.commit()
        flash('VLAN creada exitosamente', 'success')
        return redirect(url_for('vlans'))
    
    switches = Switch.query.filter_by(estado='Activo').all()
    return render_template('vlans_form.html', vlan=None, switches=switches)

@app.route('/vlans/<int:id>')
@login_required
def vlans_detalle(id):
    """Ver detalle de una VLAN"""
    vlan = VLAN.query.get_or_404(id)
    return render_template('vlans_detalle.html', vlan=vlan)

@app.route('/vlans/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def vlans_editar(id):
    """Editar una VLAN existente"""
    vlan = VLAN.query.get_or_404(id)
    
    if request.method == 'POST':
        vlan.vlan_id = int(request.form['vlan_id'])
        vlan.vlan_nombre = request.form['vlan_nombre']
        vlan.vlan_descripcion = request.form.get('vlan_descripcion')
        vlan.red = request.form.get('red')
        vlan.mascara = request.form.get('mascara')
        vlan.gateway = request.form.get('gateway')
        vlan.switch_id = request.form.get('switch_id') or None
        vlan.estado = request.form.get('estado', 'Activo')
        vlan.fecha_creacion = datetime.strptime(request.form['fecha_creacion'], '%Y-%m-%d').date() if request.form.get('fecha_creacion') else None
        vlan.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('VLAN actualizada exitosamente', 'success')
        return redirect(url_for('vlans_detalle', id=id))
    
    switches = Switch.query.filter_by(estado='Activo').all()
    return render_template('vlans_form.html', vlan=vlan, switches=switches)

@app.route('/vlans/<int:id>/eliminar', methods=['POST'])
@login_required
def vlans_eliminar(id):
    """Eliminar una VLAN"""
    vlan = VLAN.query.get_or_404(id)
    db.session.delete(vlan)
    db.session.commit()
    flash('VLAN eliminada exitosamente', 'success')
    return redirect(url_for('vlans'))

# ============================================================================
# GESTIÓN DE TOPOLOGÍA DE RED - CONEXIONES BOCA A BOCA
# ============================================================================

def obtener_conexiones_switch(switch_id):
    """Obtiene todas las conexiones de un switch específico"""
    conexiones = ConexionTopologia.query.filter(
        or_(
            and_(ConexionTopologia.equipo_origen_tipo == 'Switch', ConexionTopologia.equipo_origen_id == switch_id),
            and_(ConexionTopologia.equipo_destino_tipo == 'Switch', ConexionTopologia.equipo_destino_id == switch_id)
        )
    ).all()
    
    conexiones_detalladas = []
    for conn in conexiones:
        detalle = {
            'conexion': conn,
            'es_origen': conn.equipo_origen_tipo == 'Switch' and conn.equipo_origen_id == switch_id,
            'otro_equipo': None,
            'otro_equipo_tipo': None
        }
        
        # Determinar el otro equipo en la conexión
        if detalle['es_origen']:
            if conn.equipo_destino_tipo == 'Switch':
                detalle['otro_equipo'] = Switch.query.get(conn.equipo_destino_id)
                detalle['otro_equipo_tipo'] = 'Switch'
            elif conn.equipo_destino_tipo == 'Camara':
                detalle['otro_equipo'] = Camara.query.get(conn.equipo_destino_id)
                detalle['otro_equipo_tipo'] = 'Cámara'
            elif conn.equipo_destino_tipo == 'NVR':
                detalle['otro_equipo'] = NVR_DVR.query.get(conn.equipo_destino_id)
                detalle['otro_equipo_tipo'] = 'NVR/DVR'
        else:
            if conn.equipo_origen_tipo == 'Switch':
                detalle['otro_equipo'] = Switch.query.get(conn.equipo_origen_id)
                detalle['otro_equipo_tipo'] = 'Switch'
            elif conn.equipo_origen_tipo == 'Camara':
                detalle['otro_equipo'] = Camara.query.get(conn.equipo_origen_id)
                detalle['otro_equipo_tipo'] = 'Cámara'
            elif conn.equipo_origen_tipo == 'NVR':
                detalle['otro_equipo'] = NVR_DVR.query.get(conn.equipo_origen_id)
                detalle['otro_equipo_tipo'] = 'NVR/DVR'
        
        conexiones_detalladas.append(detalle)
    
    return conexiones_detalladas

def calcular_topologia_gabinete(gabinete_id, max_profundidad=10):
    """Calcula la topología completa hasta un gabinete específico"""
    gabinete = Gabinete.query.get_or_404(gabinete_id)
    
    topologia = {
        'gabinete': gabinete,
        'equipos_dentro': [],
        'conexiones_externas': [],
        'ruta_completa': []
    }
    
    # Equipos dentro del gabinete
    switches = Switch.query.filter_by(gabinete_id=gabinete_id, estado='Activo').all()
    nvr_dvrs = NVR_DVR.query.filter_by(gabinete_id=gabinete_id, estado='Activo').all()
    ups = UPS.query.filter_by(gabinete_id=gabinete_id, estado='Activo').all()
    fuentes = Fuente_Poder.query.filter_by(gabinete_id=gabinete_id, estado='Activo').all()
    camaras = Camara.query.filter_by(gabinete_id=gabinete_id, estado='Activo').all()
    
    topologia['equipos_dentro'] = {
        'switches': switches,
        'nvr_dvrs': nvr_dvrs,
        'ups': ups,
        'fuentes_poder': fuentes,
        'camaras': camaras
    }
    
    # Conexiones externas del gabinete
    for switch in switches:
        conexiones = obtener_conexiones_switch(switch.id)
        for conn in conexiones:
            if conn['otro_equipo']:
                topologia['conexiones_externas'].append({
                    'switch': switch,
                    'conexion': conn['conexion'],
                    'otro_equipo': conn['otro_equipo'],
                    'otro_equipo_tipo': conn['otro_equipo_tipo']
                })
    
    # Construir ruta desde Core
    topologia['ruta_completa'] = construir_ruta_hacia_core(switches)
    
    return topologia

def construir_ruta_hacia_core(switches):
    """Construye la ruta desde los switches hacia el core"""
    ruta = []
    
    # Esto es una versión simplificada - en un sistema real necesitarías 
    # una jerarquía más compleja de switches
    for switch in switches:
        if switch.gabinete_id:  # Tiene gabinete asignado
            gabinete = Gabinete.query.get(switch.gabinete_id)
            if gabinete:
                ruta.append({
                    'nivel': 1,
                    'tipo': 'Gabinete',
                    'equipo': gabinete,
                    'descripcion': f"Gabinete {gabinete.codigo} - {gabinete.nombre or ''}"
                })
            
            ruta.append({
                'nivel': 2,
                'tipo': 'Switch',
                'equipo': switch,
                'descripcion': f"Switch {switch.codigo} - {switch.ip or 'Sin IP'}"
            })
            
            # Cámaras conectadas
            camaras = Camara.query.filter_by(switch_id=switch.id, estado='Activo').all()
            for camara in camaras:
                ruta.append({
                    'nivel': 3,
                    'tipo': 'Cámara',
                    'equipo': camara,
                    'descripcion': f"Cámara {camara.codigo} - {camara.ip or 'Sin IP'}"
                })
            
            # NVR/DVR conectados
            for puerto in switch.puertos:
                if puerto.nvr_id:
                    nvr = NVR_DVR.query.get(puerto.nvr_id)
                    if nvr:
                        ruta.append({
                            'nivel': 3,
                            'tipo': 'NVR/DVR',
                            'equipo': nvr,
                            'descripcion': f"NVR {nvr.codigo} - {nvr.ip or 'Sin IP'}"
                        })
    
    return ruta

# Rutas para gestión de topología
@app.route('/topologia/switch/<int:switch_id>')
@login_required
def topologia_switch(switch_id):
    """Muestra la topología de conexiones de un switch específico"""
    switch = Switch.query.get_or_404(switch_id)
    conexiones = obtener_conexiones_switch(switch_id)
    
    # Estadísticas de conexiones
    stats = {
        'total_conexiones': len(conexiones),
        'switches_conectados': sum(1 for c in conexiones if c['otro_equipo_tipo'] == 'Switch'),
        'camaras_conectadas': sum(1 for c in conexiones if c['otro_equipo_tipo'] == 'Cámara'),
        'nvr_conectados': sum(1 for c in conexiones if c['otro_equipo_tipo'] == 'NVR/DVR')
    }
    
    return render_template('topologia_switch.html', 
                         switch=switch, 
                         conexiones=conexiones, 
                         stats=stats)

@app.route('/topologia/gabinete/<int:gabinete_id>')
@login_required
def topologia_gabinete(gabinete_id):
    """Muestra la topología completa hasta un gabinete específico"""
    topologia = calcular_topologia_gabinete(gabinete_id)
    
    # Contar equipos
    stats = {
        'switches': len(topologia['equipos_dentro']['switches']),
        'nvr_dvrs': len(topologia['equipos_dentro']['nvr_dvrs']),
        'ups': len(topologia['equipos_dentro']['ups']),
        'fuentes_poder': len(topologia['equipos_dentro']['fuentes_poder']),
        'camaras': len(topologia['equipos_dentro']['camaras']),
        'conexiones_externas': len(topologia['conexiones_externas'])
    }
    
    return render_template('topologia_gabinete.html', 
                         topologia=topologia, 
                         stats=stats)

@app.route('/conexiones')
@login_required
def conexiones_lista():
    """Lista todas las conexiones de topología"""
    conexiones = ConexionTopologia.query.order_by(ConexionTopologia.fecha_conexion.desc()).all()
    
    # Agrupar por tipo de conexión
    conexiones_agrupadas = {
        'UTP': [],
        'FibraOptica': [],
        'EnlaceInalambrico': [],
        'PoE': []
    }
    
    for conn in conexiones:
        tipo = conn.tipo_conexion
        if tipo in conexiones_agrupadas:
            conexiones_agrupadas[tipo].append(conn)
    
    return render_template('conexiones_lista.html', 
                         conexiones=conexiones,
                         conexiones_agrupadas=conexiones_agrupadas)

# API para obtener detalle de conexión
@app.route('/api/conexiones/<int:conexion_id>/detalle')
@login_required
def api_conexion_detalle(conexion_id):
    """API para obtener detalle de una conexión específica"""
    try:
        conexion = ConexionTopologia.query.get_or_404(conexion_id)
        
        html = f"""
        <div class="row">
            <div class="col-md-6">
                <h6 class="text-success">Equipo Origen</h6>
                <p><strong>Tipo:</strong> {conexion.equipo_origen_tipo}</p>
                <p><strong>Equipo:</strong> {obtener_codigo_equipo(conexion.equipo_origen_tipo, conexion.equipo_origen_id)}</p>
                <p><strong>Puerto:</strong> {conexion.puerto_origen or 'No especificado'}</p>
            </div>
            <div class="col-md-6">
                <h6 class="text-danger">Equipo Destino</h6>
                <p><strong>Tipo:</strong> {conexion.equipo_destino_tipo}</p>
                <p><strong>Equipo:</strong> {obtener_codigo_equipo(conexion.equipo_destino_tipo, conexion.equipo_destino_id)}</p>
                <p><strong>Puerto:</strong> {conexion.puerto_destino or 'No especificado'}</p>
            </div>
        </div>
        <hr>
        <div class="row">
            <div class="col-md-6">
                <p><strong>Tipo de Conexión:</strong> {conexion.tipo_conexion}</p>
                <p><strong>Velocidad:</strong> {conexion.velocidad_conexion or 'No especificada'}</p>
                <p><strong>Distancia:</strong> {conexion.distancia_metros or 0} metros</p>
            </div>
            <div class="col-md-6">
                <p><strong>Estado:</strong> <span class="badge badge-{'success' if conexion.estado_conexion == 'Activo' else 'secondary' if conexion.estado_conexion == 'Inactivo' else 'danger'}">{conexion.estado_conexion}</span></p>
                <p><strong>Fecha Conexión:</strong> {conexion.fecha_conexion.strftime('%d/%m/%Y') if conexion.fecha_conexion else 'No especificada'}</p>
            </div>
        </div>
        """
        
        if conexion.observaciones:
            html += f"""
            <hr>
            <h6>Observaciones:</h6>
            <p>{conexion.observaciones}</p>
            """
        
        return jsonify({'html': html})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/conexiones/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'superadmin')
def conexiones_editar(id):
    """Editar una conexión existente"""
    conexion = ConexionTopologia.query.get_or_404(id)
    
    if request.method == 'POST':
        # Lógica de edición similar a la creación
        conexion.equipo_origen_tipo = request.form['equipo_origen_tipo']
        conexion.equipo_origen_id = int(request.form['equipo_origen_id'])
        conexion.puerto_origen = request.form.get('puerto_origen')
        conexion.equipo_destino_tipo = request.form['equipo_destino_tipo']
        conexion.equipo_destino_id = int(request.form['equipo_destino_id'])
        conexion.puerto_destino = request.form.get('puerto_destino')
        conexion.tipo_conexion = request.form['tipo_conexion']
        conexion.distancia_metros = int(request.form.get('distancia_metros', 0)) if request.form.get('distancia_metros') else None
        conexion.velocidad_conexion = request.form.get('velocidad_conexion')
        conexion.estado_conexion = request.form.get('estado_conexion', 'Activo')
        conexion.origen_ubicacion_id = int(request.form['origen_ubicacion_id']) if request.form.get('origen_ubicacion_id') else None
        conexion.destino_ubicacion_id = int(request.form['destino_ubicacion_id']) if request.form.get('destino_ubicacion_id') else None
        conexion.observaciones = request.form.get('observaciones')
        
        db.session.commit()
        flash('Conexión actualizada exitosamente', 'success')
        return redirect(url_for('conexiones_lista'))
    
    # Para GET: obtener datos para formularios
    switches = Switch.query.filter_by(estado='Activo').all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    camaras = Camara.query.filter_by(estado='Activo').all()
    nvr_dvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    
    return render_template('conexiones_form.html', conexion=conexion,
                         switches=switches, gabinetes=gabinetes,
                         camaras=camaras, nvr_dvrs=nvr_dvrs, ubicaciones=ubicaciones)

@app.route('/conexiones/<int:id>/eliminar', methods=['POST'])
@login_required
@role_required('admin', 'superadmin')
def conexiones_eliminar(id):
    """Eliminar una conexión"""
    conexion = ConexionTopologia.query.get_or_404(id)
    db.session.delete(conexion)
    db.session.commit()
    flash('Conexión eliminada exitosamente', 'success')
    return redirect(url_for('conexiones_lista'))

@app.route('/conexiones/nueva', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'superadmin')
def conexiones_nueva():
    """Crear nueva conexión de topología"""
    if request.method == 'POST':
        tipo_conexion = request.form['tipo_conexion']
        
        nueva_conexion = ConexionTopologia(
            equipo_origen_tipo=request.form['equipo_origen_tipo'],
            equipo_origen_id=int(request.form['equipo_origen_id']),
            puerto_origen=request.form.get('puerto_origen'),
            equipo_destino_tipo=request.form['equipo_destino_tipo'],
            equipo_destino_id=int(request.form['equipo_destino_id']),
            puerto_destino=request.form.get('puerto_destino'),
            tipo_conexion=tipo_conexion,
            distancia_metros=int(request.form.get('distancia_metros', 0)) if request.form.get('distancia_metros') else None,
            velocidad_conexion=request.form.get('velocidad_conexion'),
            origen_ubicacion_id=int(request.form['origen_ubicacion_id']) if request.form.get('origen_ubicacion_id') else None,
            destino_ubicacion_id=int(request.form['destino_ubicacion_id']) if request.form.get('destino_ubicacion_id') else None,
            observaciones=request.form.get('observaciones')
        )
        
        db.session.add(nueva_conexion)
        db.session.commit()
        flash('Conexión creada exitosamente', 'success')
        return redirect(url_for('conexiones_lista'))
    
    # Obtener datos para formularios
    switches = Switch.query.filter_by(estado='Activo').all()
    gabinetes = Gabinete.query.filter_by(estado='Activo').all()
    camaras = Camara.query.filter_by(estado='Activo').all()
    nvr_dvrs = NVR_DVR.query.filter_by(estado='Activo').all()
    ubicaciones = Ubicacion.query.filter_by(activo=True).all()
    
    return render_template('conexiones_form.html',
                         switches=switches,
                         gabinetes=gabinetes,
                         camaras=camaras,
                         nvr_dvrs=nvr_dvrs,
                         ubicaciones=ubicaciones)

# ============================================================================
# DASHBOARD DE CONECTIVIDAD - Métricas de Enlaces
# ============================================================================

@app.route('/dashboard-conectividad')
@login_required
def dashboard_conectividad():
    """Dashboard con métricas de conectividad"""
    enlaces = Enlace.query.all()
    
    # Estadísticas
    total_enlaces = len(enlaces)
    enlaces_activos = sum(1 for e in enlaces if e.estado_conexion == 'Activo')
    enlaces_degradados = sum(1 for e in enlaces if e.estado_conexion == 'Degradado')
    enlaces_inactivos = sum(1 for e in enlaces if e.estado_conexion == 'Inactivo')
    
    # Promedios
    latencias = [e.latencia_ms for e in enlaces if e.latencia_ms]
    perdidas = [e.porcentaje_perdida_paquetes for e in enlaces if e.porcentaje_perdida_paquetes]
    
    latencia_promedio = sum(latencias) / len(latencias) if latencias else 0
    perdida_promedio = sum(perdidas) / len(perdidas) if perdidas else 0
    
    return render_template('dashboard_conectividad.html',
                         enlaces=enlaces,
                         total_enlaces=total_enlaces,
                         enlaces_activos=enlaces_activos,
                         enlaces_degradados=enlaces_degradados,
                         enlaces_inactivos=enlaces_inactivos,
                         latencia_promedio=round(latencia_promedio, 2),
                         perdida_promedio=round(perdida_promedio, 2))
