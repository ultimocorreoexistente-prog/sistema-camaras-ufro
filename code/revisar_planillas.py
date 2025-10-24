#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl

# Leer planilla de Fallas
print("=" * 70)
print("PLANILLA: Fallas_Actualizada.xlsx")
print("=" * 70)
try:
    wb = openpyxl.load_workbook('user_input_files/planillas-web/Fallas_Actualizada.xlsx')
    print(f"Hojas disponibles: {wb.sheetnames}")
    ws = wb.active
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    print("\nPrimeras 3 filas:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"  {i}: {row}")
        if i >= 3:
            break
except Exception as e:
    print(f"Error: {e}")

print("\n" + "=" * 70)
print("PLANILLA: Equipos_Tecnicos.xlsx")
print("=" * 70)
try:
    wb = openpyxl.load_workbook('user_input_files/planillas-web/Equipos_Tecnicos.xlsx')
    print(f"Hojas disponibles: {wb.sheetnames}")
    ws = wb.active
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    print("\nTODAS las filas:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"  {i}: {row}")
except Exception as e:
    print(f"Error: {e}")

print("\n" + "=" * 70)
print("PLANILLA: Mantenimientos.xlsx")
print("=" * 70)
try:
    wb = openpyxl.load_workbook('user_input_files/planillas-web/Mantenimientos.xlsx')
    print(f"Hojas disponibles: {wb.sheetnames}")
    ws = wb.active
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    print("\nPrimeras 3 filas:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"  {i}: {row}")
        if i >= 3:
            break
except Exception as e:
    print(f"Error: {e}")
